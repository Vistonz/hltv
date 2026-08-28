import re 
import os 
import sys
import ssl
import time

VENDOR_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".vendor")
if os.path.isdir(VENDOR_DIR) and VENDOR_DIR not in sys.path:
    sys.path.insert(0, VENDOR_DIR)

ssl._create_default_https_context = ssl._create_unverified_context

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError("Missing dependency: openpyxl. Install with: pip install -r requirements.txt") from exc

try:
    import undetected_chromedriver as uc
except ImportError as exc:
    raise ImportError("Missing dependency: undetected-chromedriver. Install with: pip install -r requirements.txt") from exc



webfront="https://www.hltv.org/stats/players"

eventfilter="?event=8876&event=8246&event=8575&event=8240&event=8047&event=8241&event=8412&event=8413&event=8248&event=8048&event=8242&event=8250&event=8049&event=8243&event=8263&event=9028&event=9029&event=8301&event=8914"
bigeventfilter = "?&event=8246&event=8575&event=8240&event=8047&event=8413&event=8248&event=8242&event=8250&event=8049&event=8243&event=8301"
eliteeventfilter = "?&event=8240&event=8876&event=8047&event=8248&event=8301&matchType=Lan"  
supereliteeventfilter = "?&event=8240&event=8876&event=8301&matchType=Lan"
arenafilter="?&event=8240&event=8876&event=8047&event=8413&event=8248&event=8242&event=8250&event=8049&event=8243&event=8301&playoffMatchType=PLAYOFFS&matchType=Lan"
keyword = ">"   
eliteeventplayofffilter = eliteeventfilter+"&playoffMatchType=PLAYOFFS"
supereliteeventplayofffilter= supereliteeventfilter+"&playoffMatchType=PLAYOFFS"
bigeventplayofffilter=bigeventfilter+"&playoffMatchType=PLAYOFFS"
bigeventfinalfilter=bigeventfilter+"&playoffMatchType=GRAND_FINAL"

minMapCountfilter="&minMapCount=65" # 图池数筛选器，更改数字使用
minrating = 1.04 #大于此数的选手rating纳入统计
file_path = os.path.join("/home/hongbin/Desktop/hltv/hltv year", "rating2026.xlsx") #表格文件保存路径，根据需要更改

wb = Workbook()
ws = wb.active
driver = uc.Chrome(version_main=149)
#用于填写表格的第一行
description = ["选手ID","图池数","rating","rounds","CTrating","Trating","首杀尝试","首杀成功","首杀rating","手枪局rating","DPR","KAST","RS","ADR","KPR","DPR_eco","KAST_eco","MULTIKILL_eco","ADR_eco","KPR_eco","KD diff","map vs top5","rating vs top5","map vs top10","rating vs top10","map vs top20","rating vs top20","回合首杀数","Rounds with a kill","ROunds with a multikill","3+kill","0.85+","1.00+","1.15+","1.30+","1.45+","clutch win","clutch point per round","big event map","big event rating","elite event map","elite event rating","superelite event map","supereliteevent rating","big event playoff map","big event playoff rating","elite event playoff map","elite event playoff rating","superelite event playoff map","superelite event playoff rating","kill per round win","adr win","win after first kill","save per round lose","assist kill percentage","damage per kill","last alive percentage","kpr lose","adr lose","traded_kill","traded_death_percentage","flash_assist","utility_damage","firepower","entrying","trading","opening","clutching","sniping","utility","HS%","singlemapwinrating","traded_death","save_teammate","saved_by_teammate","support_round_percent","attack_in_round","winrate_1v1","livetime_perround","snipkill_perround","snipkill_percent","snipkillround_percent","utlility_kill_round","throw_flash_perround","time_opponent_flashed","traded_death_percentage","assist_per_round","arena map","arena rating","big event final map","big event final rating","avg_weaponvalve_perkill"]
# 模块化
def multi_event_scraper(s):
    driver.get(s)
    time.sleep(0.6657)              
    content = driver.page_source
    namee = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>',content)
    mapp = re.findall('</span></td>\n                    <td class="statsDetail">(.*?)</td>',content)
    ratingg = re.findall('class="ratingCol(.*?)</td>',content)
    return namee,mapp,ratingg

def inner_stats_scraper(s):
    driver.get(s) 
    time.sleep(0.6657)
    content = driver.page_source
    namee = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>',content)
    ratingg = re.findall('class="ratingCol(.*?)</td>',content)
    return namee,ratingg

import unicodedata

def upload_excel1(s):
    try:
        if isinstance(s, str) and s.strip().endswith('%'):
            val = float(s.strip().replace('%', '')) / 100.0
            row.append(val)
            return
    except ValueError:
        pass

    try:
        val = float(s)
        row.append(val)
        return
    except ValueError:
        pass

    try:
        val = unicodedata.numeric(s)
        row.append(val)
        return
    except (TypeError, ValueError):
        pass

    row.append(s)

def upload_excel2(namee,ratingg):
    for name1,rating1 in zip(namee,ratingg):
        name1=name1.split(keyword, 1)[-1].strip()
        rating1=rating1.split(keyword, 1)[-1].strip()
        if name1 == name:
            upload_excel1(rating1)
            break

def upload_excel3(namee,mapp,ratingg):
    ishere = 0
    for name1,map1,rating1 in zip(namee,mapp,ratingg):
        name1=name1.split(keyword, 1)[-1].strip()
        rating1=rating1.split(keyword, 1)[-1].strip()
        if name1 == name:
            upload_excel1(map1)
            upload_excel1(rating1)
            ishere = 1
            break
    if ishere ==0:
        upload_excel1(0)
        upload_excel1("-")

# 读取现有数据并构建字典
player_data = dict()

if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):  # 从第二行开始
        name = row[0]
        id = row[1]
        player_data[name] = list(row)  # 将整行数据存储为列表
else:
    wb = Workbook()
    ws = wb.active
    ws.append(description)  # 添加表头


# 普通rating，图池
driver.get(webfront+eventfilter+minMapCountfilter) 
time.sleep(30.6657) 
content = driver.page_source 
Name = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>',content)
Mapcount = re.findall('</span></td>\n                    <td class="statsDetail">(.*?)</td>',content)
Rating = re.findall('class="ratingCol(.*?)</td>',content)
Id = re.findall('<a href="/stats/players(.*?)" data-tooltip-id="uniqueTooltipId',content)
Rounds = re.findall('<td class="statsDetail gtSmartphone-only">(.*?)</td>',content)

#名字和rating的数据清理
tempzz = 0
while tempzz < len(Name):
    Name[tempzz]=Name[tempzz].split(keyword, 1)[-1].strip()
    Rating[tempzz]=Rating[tempzz].split(keyword, 1)[-1].strip()
    tempzz = tempzz + 1

#CT,T,pistols
CTname,CTrating=inner_stats_scraper(webfront+eventfilter+"&side=COUNTER_TERRORIST"+minMapCountfilter) 
Tname,Trating=inner_stats_scraper(webfront+eventfilter+"&side=TERRORIST"+minMapCountfilter) 
Pistolname,Pistolrating =inner_stats_scraper(webfront+"/pistols"+eventfilter+minMapCountfilter)

#openingkill
driver.get(webfront+"/openingkills"+eventfilter+minMapCountfilter)
time.sleep(0.6657)
content = driver.page_source
ok = re.findall('class="statsDetail">(.*?)</td>',content)
Openingkillrating = re.findall('class="ratingCol(.*?)</td>',content)
Openingkillname = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>',content)
count=0
Openingkillattempts=[]
Openingkillsuccess=[]
for i in ok:
    if count%4 == 2:
        Openingkillattempts.append(i)
    if count%4 == 3:
        Openingkillsuccess.append(i)
    count+=1

#赛事扫描
Bigeventname,Bigeventmap,Bigeventrating = multi_event_scraper(webfront+bigeventfilter+"&minMapCount=0")
Arenaeventname,Arenaeventmap,Arenaeventrating = multi_event_scraper(webfront+arenafilter+"&minMapCount=0")
Supereliteeventname,Supereliteeventmap,Supereliteeventrating = multi_event_scraper(webfront+supereliteeventfilter+"&minMapCount=0")
Eliteeventname,Eliteeventmap,Eliteeventrating = multi_event_scraper(webfront+eliteeventfilter+"&minMapCount=0")
Eliteeventplayoffname,Eliteeventplayoffmap,Eliteeventplayoffrating = multi_event_scraper(webfront+eliteeventplayofffilter+"&minMapCount=0")
Supereliteeventplayoffname,Supereliteeventplayoffmap,Supereliteeventplayoffrating = multi_event_scraper(webfront+supereliteeventplayofffilter+"&minMapCount=0")
Bigeventplayoffname,Bigeventplayoffmap,Bigeventplayoffrating = multi_event_scraper(webfront+bigeventplayofffilter+"&minMapCount=0")
Bigeventfinalname,Bigeventfinalmap,Bigeventfinalrating = multi_event_scraper(webfront+bigeventfinalfilter+"&minMapCount=0")
'''
#Bigeventplayoff
time.sleep(0.6657)
tempzz3 = 0
Bigeventplayoffname = []
Bigeventplayoffmap = []
Bigeventplayoffratingsum = []
Bigeventplayoffrounds = []
while tempzz3 < len(Eliteeventplayoffname):
    Bigeventplayoffname.append(Eliteeventplayoffname[tempzz3])
    Bigeventplayoffmap.append(Eliteeventplayoffmap[tempzz3])
    Bigeventplayoffratingsum.append(Eliteeventplayoffratingsum[tempzz3])
    Bigeventplayoffrounds.append(Eliteeventplayoffrounds[tempzz3])
    tempzz3 = tempzz3 +1
tempzz3 = 0
for i,j in zip(bigevent,bigeventplayoff):
    driver.get(webfront+"?"+i+j)
    content = driver.page_source
    bigeventplayofftempname = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>',content)
    bigeventplayofftempmap = re.findall('<td class="statsDetail">(.*?)</td>',content)
    bigeventplayoffrating = re.findall('class="ratingCol(.*?)</td>',content)
    bigeventplayoffrounds = re.findall('<td class="statsDetail gtSmartphone-only">(.*?)</td>',content)
    bigeventplayoffmap = []
    bigeventplayoffname = []
    tempzz2 = 0
    while tempzz2 <len(bigeventplayofftempmap):
        if tempzz2 ==0 or tempzz2%2==0:
            bigeventplayoffmap.append(bigeventplayofftempmap[tempzz2])
            bigeventplayoffname.append(bigeventplayofftempname[tempzz2])
        tempzz2 =tempzz2+1
    tempzz2 = 0
    while tempzz2 < len(bigeventplayoffname):
        bigeventplayoffname[tempzz2]=bigeventplayoffname[tempzz2].split(keyword, 1)[-1].strip()
        bigeventplayoffrating[tempzz2]=bigeventplayoffrating[tempzz2].split(keyword, 1)[-1].strip()
        tempzz2 = tempzz2 + 1
    for name1,map1,rating1,round1 in zip(bigeventplayoffname,bigeventplayoffmap,bigeventplayoffrating,bigeventplayoffrounds):
        #过滤异常数据
        name1=name1.split(keyword, 1)[-1].strip()
        rating1=rating1.split(keyword, 1)[-1].strip()
        belong = 0
        tempzz1 = 0
        while tempzz1 < len(Bigeventplayoffname):
            if Bigeventplayoffname[tempzz1] == name1:                
                Bigeventplayoffmap[tempzz1] = Bigeventplayoffmap[tempzz1] + float(map1)
                Bigeventplayoffratingsum[tempzz1] = Bigeventplayoffratingsum[tempzz1] + (float(rating1) * float(round1))
                Bigeventplayoffrounds[tempzz1] = Bigeventplayoffrounds[tempzz1] +float(round1)
                belong = 1
                break
            tempzz1 = tempzz1 + 1
        if belong == 0:
            Bigeventplayoffname.append(name1)
            Bigeventplayoffmap.append(float(map1))
            Bigeventplayoffratingsum.append(float(rating1)*float(round1))
            Bigeventplayoffrounds.append(float(round1))
'''

#选手个人数据统计和打印
for name,mapcount,rating,id,rounds in zip(Name,Mapcount,Rating,Id,Rounds):
    #加入选手姓名，图池数，rating，回合数（用于计算）
    row = [name,int(mapcount),float(rating),int(rounds)]
    if name in player_data:
        old_mapcount = player_data[name][1]  # 获取之前的地图池数
        if old_mapcount == int(mapcount):
            print(f"⏩ 跳过已完成且地图池未变化的选手：{name}")
            continue  # 如果地图池数没变，跳过该选手
        else:
            print(f"🔄 更新选手 {name} 的数据，地图池变化")
    else:
        print(f"✅ 正在爬取新选手：{name}")
    ratingnum = float(rating)
    if ratingnum <= minrating:
        break
    #上传CTrating,Trating
    upload_excel2(CTname,CTrating)
    upload_excel2(Tname,Trating)    
    #上传首杀相关数据（回合首杀在选手个人页面中）
    for openingkillname,openingkillattempts,openingkillsuccess,openingkillrating in zip(Openingkillname,Openingkillattempts,Openingkillsuccess,Openingkillrating):
        openingkillname=openingkillname.split(keyword, 1)[-1].strip()
        openingkillrating=openingkillrating.split(keyword, 1)[-1].strip()
        if openingkillname == name:
            upload_excel1(openingkillattempts)
            upload_excel1(openingkillsuccess)
            upload_excel1(openingkillrating)
            break
    upload_excel2(Pistolname,Pistolrating)
    #网址清理，用于进入选手个人界面
    statsuffix = id.replace('amp;','')
    driver.get(webfront+statsuffix)
    content = driver.page_source
    time.sleep(0.6657)
    #获取选手面板数据，rating已经统计不被需要
    Playerstatform=re.findall('<div class="player-summary-stat-box-data traditionalData">(.*?)</div>',content)
    DPR = Playerstatform[0]
    upload_excel1(DPR)
    KAST=re.findall('<div class="player-summary-stat-box-data traditionalData">(.*?)<span',content)
    upload_excel1(KAST[0])
    RS = re.findall('<div class="player-summary-stat-box-data">(.*?)<span class="',content)
    upload_excel1(RS[0])
    ADR = Playerstatform[3]
    upload_excel1(ADR)
    KPR = Playerstatform[4]
    upload_excel1(KPR)
    Playerstatform_eco=re.findall('<div class="player-summary-stat-box-data ecoAdjustedData hidden">(.*?)</div>',content)
    DPRE = Playerstatform_eco[0]
    upload_excel1(DPRE)
    KASTE=re.findall('<div class="player-summary-stat-box-data ecoAdjustedData hidden">(.*?)<span',content)
    upload_excel1(KASTE[0])
    MULTIKILL = Playerstatform_eco[2]
    upload_excel1(MULTIKILL)
    ADRE = Playerstatform_eco[3]
    upload_excel1(ADRE)
    KPRE = Playerstatform_eco[4]
    upload_excel1(KPRE)
    kills = re.findall('Total kills</span><span>(.*?)</span></div>',content)
    kills = kills[0]
    deaths = re.findall('Total deaths</span><span>(.*?)</span></div>',content)
    deaths = deaths[0]
    headshot = re.findall('Headshot %</span><span>(.*?)</span></div>',content)
    upload_excel1(int(kills)-int(deaths))
    chartstat = re.findall('<div class="role-stats-data">(.*?)</div>',content)
    charttotalstat = re.findall('"row-stats-section-score">(.*?)<',content)
    #新增数据 
    saved_by_teammate = chartstat[24]

    livetime_perround_str = chartstat[84]
    livetime = re.findall(r"\d+\.?\d*",livetime_perround_str)
    livetime_perround = float(livetime[0])*60.0+float(livetime[1])

    #统计选手对阵top5数据
    VStop5maplist =   re.findall('vs top 5 opponents</div>\n                      <div class="rating-maps">(.*?)</div>',content)
    VStop5map = VStop5maplist[0]
    VStop5map = VStop5map.replace('(','')
    VStop5map = VStop5map.replace(' maps)','')
    upload_excel1(VStop5map)
    VStop5rating = re.findall('<div class="rating-value">(.*?)</div>\n                      <div class="rating-description">vs top 5 opponents',content)
    upload_excel1(VStop5rating[0])
    #统计选手对阵top10数据
    VStop10maplist = re.findall('vs top 10 opponents</div>\n                      <div class="rating-maps">(.*?)</div>',content)
    VStop10map = VStop10maplist[0]
    VStop10map = VStop10map.replace('(','')
    VStop10map = VStop10map.replace(' maps)','')
    upload_excel1(VStop10map)
    VStop10rating = re.findall('<div class="rating-value">(.*?)</div>\n                      <div class="rating-description">vs top 10 opponents',content)
    upload_excel1(VStop10rating[0])
    #统计选手对阵top20数据
    VStop20maplist = re.findall('vs top 20 opponents</div>\n                      <div class="rating-maps">(.*?)</div>',content)
    VStop20map = VStop20maplist[0]
    VStop20map = VStop20map.replace('(','')
    VStop20map = VStop20map.replace(' maps)','')
    upload_excel1(VStop20map)
    VStop20rating = re.findall('<div class="rating-value">(.*?)</div>\n                      <div class="rating-description">vs top 20 opponents',content)
    upload_excel1(VStop20rating[0])
    #进入individual界面，统计首杀数和回合击杀
    time.sleep(0.6657)
    driver.get(webfront+"/individual"+statsuffix)
    content = driver.page_source
    openingkills = re.findall('Total opening kills</span><span>(.*?)</span></div>',content)
    upload_excel1(float(openingkills[0]) / float(rounds))
    round0kill = re.findall('0 kill rounds</span><span>(.*?)</span></div>',content)
    upload_excel1((float(rounds)-float(round0kill[0]))/float(rounds))
    round1kill = re.findall('1 kill rounds</span><span>(.*?)</span></div>',content)
    round2kill = re.findall('2 kill rounds</span><span>(.*?)</span></div>',content)
    round3kill = re.findall('3 kill rounds</span><span>(.*?)</span></div>',content)
    round4kill = re.findall('4 kill rounds</span><span>(.*?)</span></div>',content)
    round5kill = re.findall('5 kill rounds</span><span>(.*?)</span></div>',content)
    upload_excel1((float(rounds)-float(round0kill[0])-float(round1kill[0]))/float(rounds))
    upload_excel1((float(round5kill[0])+float(round4kill[0])+float(round3kill[0]))/float(rounds))
    #进入match界面，统计对局的小局rating,获胜回合数
    singlemap = int(mapcount)
    offset1 = 0
    singlemapwinround = 0
    singlemapwinrating = 0
    winround = 0.0
    #统计大于指定数字rating比例
    ratingabove085 = 0
    ratingabove100 = 0
    ratingabove115 = 0
    ratingabove130 = 0
    ratingabove145 = 0
    while singlemap > 0:    
        time.sleep(0.6657)
        driver.get(webfront+"/matches"+statsuffix+"&offset="+str(offset1))
        content = driver.page_source
        singlemaprating = re.findall('"none">(.*?)</td>\n                  </tr>',content)
        mapround = re.findall(r'</span></a><span> \((.*?)\)</span></div>',content)
        tempzz7 = 0 
        #求失败回合KPR和ADR
        while tempzz7 < len(mapround)-1:
            winround += float(mapround[tempzz7])
            tempzz7+=2
        tempzz4 = 0
        singlemapround = re.findall(r'</span></a><span> \((.*?)\)</span></div>',content)
        for i in singlemaprating:
            rating1 = float(i)
            if float(singlemapround[2*tempzz4]) > float(singlemapround[2*tempzz4+1]) :
                totalround = float(singlemapround[2*tempzz4]) + float(singlemapround[2*tempzz4+1])
                singlemapwinround = singlemapwinround + totalround 
                singlemapwinrating = singlemapwinrating + (totalround * rating1)
            if rating1 >= 0.85:
                ratingabove085 += 1
            if rating1 >= 1:
                ratingabove100 += 1
            if rating1 >= 1.15:
                ratingabove115 += 1
            if rating1 >= 1.30:
                ratingabove130 += 1 
            if rating1 >= 1.45:
                ratingabove145 += 1 
            tempzz4 = tempzz4 + 1
        singlemap -= 100
        offset1 += 100
    kprl = (float(kills)-float(chartstat[6])*winround)/float(float(rounds)-winround)
    adrl = (float(float(ADR)*float(rounds))-float(chartstat[18])*winround)/float(float(rounds)-winround)
    upload_excel1(ratingabove085)
    upload_excel1(ratingabove100)
    upload_excel1(ratingabove115)
    upload_excel1(ratingabove130)
    upload_excel1(ratingabove145)

    #统计选手残局获胜次数
    pattern = r"(/[0-9]+)"
    driver.get(webfront+"/clutches"+re.sub(pattern,r'\1'+"/1on1",statsuffix,1))
    time.sleep(0.6657)
    content = driver.page_source
    clutch1v1win = re.findall('<div class="value">(.*?)</div>',content)
    driver.get(webfront+"/clutches"+re.sub(pattern,r'\1'+"/1on2",statsuffix,1))
    time.sleep(0.6657)
    content = driver.page_source
    clutch1v2win = re.findall('<div class="value">(.*?)</div>',content)
    driver.get(webfront+"/clutches"+re.sub(pattern,r'\1'+"/1on3",statsuffix,1))
    time.sleep(0.6657)
    content = driver.page_source
    clutch1v3win = re.findall('<div class="value">(.*?)</div>',content)
    driver.get(webfront+"/clutches"+re.sub(pattern,r'\1'+"/1on4",statsuffix,1))
    time.sleep(0.6657)
    content = driver.page_source
    clutch1v4win = re.findall('<div class="value">(.*?)</div>',content)
    driver.get(webfront+"/clutches"+re.sub(pattern,r'\1'+"/1on5",statsuffix,1))
    time.sleep(0.6657)
    content = driver.page_source
    clutch1v5win = re.findall('<div class="value">(.*?)</div>',content)
    if clutch1v5win[0] =="-":
        clutch1v5win[0] = "0"
    if clutch1v4win[0] =="-":
        clutch1v4win[0] = "0"
    if clutch1v3win[0] =="-":
        clutch1v3win[0] = "0"
    if clutch1v2win[0] =="-":
        clutch1v2win[0] = "0"
    if clutch1v1win[0] =="-":
        clutch1v1win[0] = "0"
    upload_excel1(int(clutch1v5win[0])+int(clutch1v4win[0])+int(clutch1v3win[0])+int(clutch1v2win[0])+int(clutch1v1win[0]))
    upload_excel1( (float(clutch1v5win[0])*16.0+float(clutch1v4win[0])*8.0+float(clutch1v3win[0])*4.0+float(clutch1v2win[0])*2.0+float(clutch1v1win[0])) /float(rounds))
    #上传bigevent等数据
    upload_excel3(Bigeventname,Bigeventmap,Bigeventrating)
    upload_excel3(Eliteeventname,Eliteeventmap,Eliteeventrating)
    upload_excel3(Supereliteeventname,Supereliteeventmap,Supereliteeventrating)
    upload_excel3(Bigeventplayoffname,Bigeventplayoffmap,Bigeventplayoffrating)
    upload_excel3(Eliteeventplayoffname,Eliteeventplayoffmap,Eliteeventplayoffrating)
    upload_excel3(Supereliteeventplayoffname,Supereliteeventplayoffmap,Supereliteeventplayoffrating)  
    upload_excel1(chartstat[6])#kprw
    upload_excel1(chartstat[18])#adrw
    upload_excel1(chartstat[69])#winafteropeningkill
    upload_excel1(chartstat[87])#saveperroundlose
    upload_excel1(chartstat[51])#assist_kill_percentage
    upload_excel1(chartstat[54])#damageperkill
    upload_excel1(chartstat[78])#last_alive_percentage
    upload_excel1(kprl)
    upload_excel1(adrl)
    upload_excel1(chartstat[45])#tradedkill
    upload_excel1(chartstat[30])#traded-death-percentage
    upload_excel1(chartstat[114])#flash-assist
    upload_excel1(chartstat[105])#utility-damage
    upload_excel1(charttotalstat[0]) #firepower等七项数据
    upload_excel1(charttotalstat[3]) 
    upload_excel1(charttotalstat[6])
    upload_excel1(charttotalstat[9])
    upload_excel1(charttotalstat[12])
    upload_excel1(charttotalstat[15])
    upload_excel1(charttotalstat[18])
    upload_excel1(headshot[0])
    if float(singlemapwinround)>0:
        upload_excel1(float(singlemapwinrating)/float(singlemapwinround))
    else:
        upload_excel1("-")
    upload_excel1(chartstat[27])#traded-death
    upload_excel1(chartstat[42])#save-teammate
    upload_excel1(chartstat[24])#save-by-teammate
    upload_excel1(chartstat[39])#support-round-percent
    upload_excel1(chartstat[72])#attack-in-round
    upload_excel1(chartstat[81])  #winrate1v1
    upload_excel1(livetime_perround)
    upload_excel1(chartstat[90])#snipkill-perround
    upload_excel1(chartstat[93])#snipkill-percent
    upload_excel1(chartstat[96])#snipkillround_percent
    upload_excel1(chartstat[108])#utility-kill-per100round
    upload_excel1(chartstat[111])#throw_flash_perround
    upload_excel1(chartstat[117])#time_opponent_flashed
    upload_excel1(chartstat[48])#tradedkillpercentage
    upload_excel1(chartstat[36])#assist-per-round
    upload_excel3(Arenaeventname,Arenaeventmap,Arenaeventrating)
    upload_excel3(Bigeventfinalname,Bigeventfinalmap,Bigeventfinalrating)

    #进入武器界面，统计选手击杀使用武器平均价值
    driver.get(webfront+"/weapon"+statsuffix)
    content = driver.page_source
    weapon_name = re.findall('.</span><span> (.*?)<',content)
    weapon_kill = re.findall('</span></div>\n<span>(.*?)<',content)
    economy_sum = 0
    kill_sum = 0
    tempzz1 = 0
    for i in weapon_name:
        if i == "awp":
            economy_sum += 4750 * float(weapon_kill[tempzz1])
        elif i == "ak47":
            economy_sum += 2700 * float(weapon_kill[tempzz1])
        elif i == "m4a1" or i == "m4a1_silencer":
            economy_sum += 2900 * float(weapon_kill[tempzz1])
        elif i == "ssg08" or i == "negev":
            economy_sum += 1700 * float(weapon_kill[tempzz1])
        elif i == "usp_silencer" or i == "hkp2000" or i == "glock" or i == "usp_silencer_off" or i == "taser":
            economy_sum += 200 * float(weapon_kill[tempzz1])
        elif i == "deagle":
            economy_sum += 700 * float(weapon_kill[tempzz1])
        elif i == "fiveseven" or i == "tec9" or i == "cz75a":
            economy_sum += 500 * float(weapon_kill[tempzz1])
        elif i ==  "revolver":
            economy_sum += 600 * float(weapon_kill[tempzz1])
        elif i == "elite" or i == "p250":
            economy_sum += 300 * float(weapon_kill[tempzz1])
        elif i == "ump45" or i == "mp9":
            economy_sum += 1250 * float(weapon_kill[tempzz1])
        elif i == "bizon":
            economy_sum += 1400 * float(weapon_kill[tempzz1])
        elif i == "mp7" or i == "mp5sd":
            economy_sum += 1500 * float(weapon_kill[tempzz1])
        elif i == "mag7":
            economy_sum += 1300 * float(weapon_kill[tempzz1])
        elif i == "galilar":
            economy_sum += 1800 * float(weapon_kill[tempzz1])
        elif i == "famas":
            economy_sum += 1950 * float(weapon_kill[tempzz1])
        elif i == "aug":
            economy_sum += 3300 * float(weapon_kill[tempzz1])
        elif i == "sg556":
            economy_sum += 3000 * float(weapon_kill[tempzz1])
        elif i == "xm1014":
            economy_sum += 2000 * float(weapon_kill[tempzz1])
        elif i == "mac10" or i == "nova":
            economy_sum += 1050 * float(weapon_kill[tempzz1])
        elif i == "g3sg1" or i == "scar20" or i == "m249":
            economy_sum += 2300 * float(weapon_kill[tempzz1])
        else:
            economy_sum += 0 * float(weapon_kill[tempzz1])
        kill_sum += float(weapon_kill[tempzz1])
        tempzz1+=1
    upload_excel1(economy_sum/kill_sum)
    print(row) #用于测试
    player_data[name] = row
    row_idx = next((i for i, r in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True)) if r[0] == name), None)
    if row_idx is not None:
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_idx + 2, column=col, value=value)
    else:
        ws.append(row)
    wb.save(file_path)
# 写入排序后的数据（按最新爬取顺序）
ws.delete_rows(2, ws.max_row)  # 删除旧数据（保留表头）

for name in Name:
    if name in player_data:
        ws.append(player_data[name])  # 按最新顺序写入表格

wb.save(file_path)
print("✅ 按新排名顺序更新 Excel 完成。")

# 关闭浏览器 
driver.close()            
driver.quit() 