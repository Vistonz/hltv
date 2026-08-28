"""
修复 player_mvp_detailed.xlsx 中 Playoffs 数据的脚本。
问题：之前爬取时 playoffMatchType=playoffs (小写)，应该是 playoffMatchType=PLAYOFFS (大写)。
此脚本重新爬取正确的 Playoffs 数据并更新 Excel 文件中的列 F-I (第6-9列)。
"""

import re
import time
from openpyxl import load_workbook
import undetected_chromedriver as uc

# URL 与工作表名的直接对应关系（按顺序一一对应）
# 工作表名是在 testMVP.py 中根据爬到的玩家名生成的
URL_TO_SHEET = [
    ("https://www.hltv.org/stats/players/3741/niko?event=8246", "NiKo_8246"),
    ("https://www.hltv.org/stats/players/22929/nota?event=8246", "nota_8246"),
    ("https://www.hltv.org/stats/players/11893/zywoo?event=8042", "ZywOo_8042"),
    ("https://www.hltv.org/stats/players/11816/ropz?event=8042", "ropz_8042"),
    ("https://www.hltv.org/stats/players/20447/heavygod?event=7912&event=7907", "HeavyGod_7912"),
    ("https://www.hltv.org/stats/players/11893/zywoo?event=7912&event=7907", "ZywOo_7912"),
    ("https://www.hltv.org/stats/players/21062/matys?event=7912&event=7907", "MATYS_7912"),
    ("https://www.hltv.org/stats/players/17306/degster?event=8044", "degster_8044"),
    ("https://www.hltv.org/stats/players/19230/m0nesy?event=8044", "m0NESY_8044"),
    ("https://www.hltv.org/stats/players/3741/niko?event=8044", "NiKo_8044"),
    ("https://www.hltv.org/stats/players/12018/teses?event=8044", "TeSeS_8044"),
    ("https://www.hltv.org/stats/players/20447/heavygod?event=8044", "HeavyGod_8044"),
    ("https://www.hltv.org/stats/players/8183/rain?event=6865", "rain_6865"),
    ("https://www.hltv.org/stats/players/11816/ropz?event=6865", "ropz_6865"),
    ("https://www.hltv.org/stats/players/18053/broky?event=6865", "broky_6865"),
    ("https://www.hltv.org/stats/players/16920/sh1ro?event=5552", "sh1ro_5552"),
    ("https://www.hltv.org/stats/players/13915/yekindar?event=5552", "YEKINDAR_5552"),
    ("https://www.hltv.org/stats/players/17372/fang?event=5454", "FaNg_5454"),
    ("https://www.hltv.org/stats/players/7938/xantares?event=5454", "XANTARES_5454"),
    ("https://www.hltv.org/stats/players/4954/xyp9x?event=3883", "Xyp9x_3883"),
    ("https://www.hltv.org/stats/players/7998/s1mple?event=3883", "s1mple_3883"),
    ("https://www.hltv.org/stats/players/7592/device?event=3883", "device_3883"),
    ("https://www.hltv.org/stats/players/9032/magisk?event=3883", "Magisk_3883"),
    ("https://www.hltv.org/stats/players/7156/msl?event=3389", "MSL_3389"),
    ("https://www.hltv.org/stats/players/7592/device?event=3389", "device_3389"),
    ("https://www.hltv.org/stats/players/8095/aizy?event=3389", "aizy_3389"),
    ("https://www.hltv.org/stats/players/9031/valde?event=3389", "valde_3389"),
    ("https://www.hltv.org/stats/players/7398/dupreeh?event=3373", "dupreeh_3373"),
    ("https://www.hltv.org/stats/players/7592/device?event=3373", "device_3373"),
    ("https://www.hltv.org/stats/players/7412/gla1ve?event=3373", "gla1ve_3373"),
    ("https://www.hltv.org/stats/players/9078/k0nfig?event=2410", "k0nfig_2410"),
    ("https://www.hltv.org/stats/players/9032/magisk?event=2410", "Magisk_2410"),
    ("https://www.hltv.org/stats/players/7148/friberg?event=1444", "friberg_1444"),
    ("https://www.hltv.org/stats/players/3055/flusha?event=1444", "flusha_1444"),
    ("https://www.hltv.org/stats/players/885/olofmeister?event=1444", "olofmeister_1444"),
    ("https://www.hltv.org/stats/players/7322/apex?event=1444", "apEX_1444"),
    ("https://www.hltv.org/stats/players/7998/s1mple?event=6140", "s1mple_6140"),
    ("https://www.hltv.org/stats/players/8183/rain?event=6140", "rain_6140"),
    ("https://www.hltv.org/stats/players/10394/twistzz?event=6140", "Twistzz_6140")
]

EXCEL_PATH = r"C:\Users\10725\Desktop\hltv\test code\player_mvp_detailed.xlsx"


def get_playoffs_stats(driver, base_url):
    """
    使用正确的 PLAYOFFS (大写) 参数爬取数据。
    返回: (maps, rating, kprw, adrw)
    """
    url = base_url + "&playoffMatchType=PLAYOFFS"
    
    driver.get(url)
    time.sleep(2.5)
    content = driver.page_source
    
    # Maps Played
    maps_match = re.search(r'<span>Maps played</span>\s*<span[^>]*>(\d+)</span>', content)
    maps = int(maps_match.group(1)) if maps_match else 0
    
    if maps == 0:
        return 0, "-", "-", "-"

    # Rating
    primary_stats = re.findall(r'class="player-summary-stat-box-rating-data-text">(.*?)</div>', content)
    rating = primary_stats[0] if primary_stats else "-"
    
    # KPRW and ADRW
    chart_stats = re.findall(r'<div class="role-stats-data">(.*?)</div>', content)
    
    kprw = "-"
    adrw = "-"
    
    if len(chart_stats) > 18:
        kprw = chart_stats[6]
        adrw = chart_stats[18]
    
    return maps, rating, kprw, adrw


def main():
    print("=" * 60)
    print("修复 Playoffs 数据脚本")
    print("=" * 60)
    
    # 加载 Excel 文件
    wb = load_workbook(EXCEL_PATH)
    print(f"已加载文件: {EXCEL_PATH}")
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"待处理 URL 数量: {len(URL_TO_SHEET)}")
    
    # 验证工作表存在
    missing_sheets = [s for _, s in URL_TO_SHEET if s not in wb.sheetnames]
    if missing_sheets:
        print(f"[警告] 以下工作表不存在: {missing_sheets}")
    
    # 启动浏览器
    driver = uc.Chrome()
    
    try:
        updated_count = 0
        
        for idx, (url, sheet_name) in enumerate(URL_TO_SHEET):
            if sheet_name not in wb.sheetnames:
                print(f"[{idx+1}/{len(URL_TO_SHEET)}] 跳过 - 工作表不存在: {sheet_name}")
                continue
            
            ws = wb[sheet_name]
            print(f"\n[{idx+1}/{len(URL_TO_SHEET)}] 处理工作表: {sheet_name}")
            
            # 显示旧数据 (列 6-9, Row 3)
            old_data = [ws.cell(row=3, column=c).value for c in range(6, 10)]
            print(f"  旧 Playoffs 数据: Maps={old_data[0]}, Rating={old_data[1]}, KPRW={old_data[2]}, ADRW={old_data[3]}")
            
            # 爬取新数据
            new_data = get_playoffs_stats(driver, url)
            print(f"  新 Playoffs 数据: Maps={new_data[0]}, Rating={new_data[1]}, KPRW={new_data[2]}, ADRW={new_data[3]}")
            
            # 更新单元格 (列 F=6, G=7, H=8, I=9, Row 3)
            for col_idx, value in enumerate(new_data, start=6):
                ws.cell(row=3, column=col_idx, value=value)
            
            updated_count += 1
            
            # 定期保存以防中断
            if updated_count % 5 == 0:
                wb.save(EXCEL_PATH)
                print(f"  [已保存进度: {updated_count} 个工作表已更新]")
        
        # 最终保存
        wb.save(EXCEL_PATH)
        print("\n" + "=" * 60)
        print(f"修复完成！共更新 {updated_count} 个工作表的 Playoffs 数据。")
        print(f"文件已保存: {EXCEL_PATH}")
        print("=" * 60)
        
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
