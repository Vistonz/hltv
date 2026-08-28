import re
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import undetected_chromedriver as uc

# ==========================================
# User Configuration: Player/Event URLs
# ==========================================
# The script will process each URL. These should be the player stats summary URLs.
PLAYER_URL_LIST = [
    "https://www.hltv.org/stats/players/11893/zywoo?event=8248",
    
]

# Output path
OUTPUT_PATH = os.path.join(r"C:\Users\10725\Desktop\hltv\test code", "player_mvp_detailed.xlsx")

# ==========================================
# Scraping Logic
# ==========================================

def get_stats_from_summary(driver, base_url, match_type=None):
    """
    Scrapes Maps, Rating, KPRW, and ADRW from the player summary page.
    Follows original code logic for role-stats.
    """
    url = base_url
    if match_type:
        url += f"&playoffMatchType={match_type}"
    
    driver.get(url)
    time.sleep(2.5)
    content = driver.page_source
    #name
    player_name = re.search(r'<div class="player-summary-stat-box-left-nickname text-ellipsis">(.*?)</div>',content)
    
    # 1. Maps Played
    maps_match = re.search(r'<span>Maps played</span>\s*<span[^>]*>(\d+)</span>', content)
    maps = int(maps_match.group(1)) if maps_match else 0
    
    if maps == 0:
        return None, 0, "-", "-", "-"

    # 2. Rating (From the primary summary box)
    primary_stats = re.findall(r'class="player-summary-stat-box-rating-data-text">(.*?)</div>', content)
    rating = primary_stats[0] if primary_stats else "-"
    
    # 3. KPRW and ADRW (From Role Stats chart)
    chart_stats = re.findall(r'<div class="role-stats-data">(.*?)</div>', content)
    
    kprw = "-"
    adrw = "-"
    
    if len(chart_stats) > 18:
        kprw = chart_stats[6]
        adrw = chart_stats[18]
    
    # Extract player name (preserve original case)
    p_name = player_name.group(1) if player_name else None
    
    return p_name, maps, rating, kprw, adrw

def get_map_rank(driver, map_url, target_player_id):
    """
    Visits the map stats page and determines player's rank (1-10) among all players.
    Identifies the player by their ID in the URL (e.g., /players/22929/nota).
    """
    if not map_url:
        return "-"
    
    full_url = "https://www.hltv.org" + map_url
    driver.get(full_url)
    time.sleep(2.0)
    content = driver.page_source
    
    # Find player stats tables specifically (class contains "totalstats")
    tables = re.findall(r'<table[^>]*class="[^"]*totalstats[^"]*"[^>]*>(.*?)</table>', content, re.DOTALL)
    
    all_ratings = []
    target_rating = None
    found_ids = []
    
    for table in tables:
        # Find all rows in this table
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL)
        for row in rows:
                
            # Extract player ID from the row
            id_match = re.search(r'/stats/players/(\d+)/', row)
            # Extract rating from the row
            rating_match = re.search(r'<td class="st-rating rating[^>]*>(.*?)</td>', row)
            
            if id_match and rating_match:
                try:
                    player_id = id_match.group(1)
                    r_val = float(rating_match.group(1).split('>')[-1].strip())
                    all_ratings.append(r_val)
                    
                    if player_id == str(target_player_id):
                        target_rating = r_val
                except: continue
        
    # Debug: show count and IDs when target not found
    if target_rating is None and all_ratings:
        # Collect IDs for debug
        found_ids = []
        for table in tables:
            rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL)
            for row in rows:
                id_match = re.search(r'/stats/players/(\d+)/', row)
                if id_match and len(found_ids) < 10:
                    found_ids.append(id_match.group(1))
        print(f"    [RANK DEBUG] Target ID {target_player_id} not in IDs: {found_ids}")
    
    if target_rating is None or not all_ratings:
        return "-"
    
    # Calculate rank
    rank = 1
    for r in all_ratings:
        if r > target_rating + 0.001:
            rank += 1
            
    return rank

def get_detailed_map_stats(driver, base_url):
    """
    Calculates weighted rating and traverses individual maps for ranks.
    """
    # Extract player ID and name from URL (e.g., .../3741/niko?...)
    id_match = re.search(r'/players/(\d+)/([^?&]+)', base_url)
    target_player_id = id_match.group(1) if id_match else "0"
    target_name = id_match.group(2) if id_match else "unknown"

    def parse_maps(content):
        # Anchor on rounds (2 entries per map) to ensure we find every map
        rounds_list = re.findall(r'</span></a><span> \((.*?)\)</span></div>', content)
        num_maps = len(rounds_list) // 2
        
        # Class-agnostic rating regex
        ratings = re.findall(r'<td class="match(.*?)</td>', content)
        # Opponent names (usually one per row next to logo)
        opps = re.findall(r'loading="lazy">(.*?)</span></a><span>', content)
        # Map stats detail URLs
        map_stats_urls = re.findall(r'<td class="no-sort" data-sort-method="none"><a href="(.*?)">', content)
        
        # Debug: Show counts for all lists
        print(f"    [DEBUG] Rounds:{len(rounds_list)}, Ratings:{len(ratings)}, Opps:{len(opps)}, URLs:{len(map_stats_urls)}")
        
        parsed = []
        for i in range(num_maps):
            try:
                r1 = float(rounds_list[2*i])
                r2 = float(rounds_list[2*i+1])
                # Extract rating value from string like '-rating-good">1.18'
                rating_str = ratings[i].split('>')[-1].strip() if i < len(ratings) else "0.0"
                rating_val = float(rating_str) if rating_str else 0.0
                opp_name = opps[2*i+1] if (2*i+1) < len(opps) else "unknown"
                m_url = map_stats_urls[i] if i < len(map_stats_urls) else None
                
                parsed.append({
                    'rating': rating_val,
                    'rounds': r1 + r2,
                    'is_win': r1 > r2,
                    'opp': opp_name,
                    'url': m_url,
                    'id': f"{opp_name}_{r1}_{r2}"
                })
            except: continue
        print(f"  Found {len(parsed)} maps.")
        return parsed

    # 1. Get ALL matches (used to find all wins)
    driver.get(base_url.replace("/players/", "/players/matches/"))
    time.sleep(2.0)
    all_matches = parse_maps(driver.page_source)
    
    # 2. Get Grand Final matches
    driver.get(base_url.replace("/players/", "/players/matches/") + "&playoffMatchType=GRAND_FINAL")
    time.sleep(2.0)
    final_matches = parse_maps(driver.page_source)

    # 3. Weighted Logic: Union = All Wins + Grand Final Losses
    win_maps = [m for m in all_matches if m['is_win']]
    final_losses = [m for m in final_matches if not m['is_win']]

    win_tr_r = sum(m['rating'] * m['rounds'] for m in win_maps)
    win_tr_w = sum(m['rounds'] for m in win_maps)
    loss_tr_r = sum(m['rating'] * m['rounds'] for m in final_losses)
    loss_tr_w = sum(m['rounds'] for m in final_losses)

    tr_r = win_tr_r + loss_tr_r
    tr = win_tr_w + loss_tr_w
    weighted_combined = round(tr_r / tr, 2) if tr > 0 else "-"
    
    win_rating_avg = round(win_tr_r / win_tr_w, 2) if win_tr_w > 0 else "-"
    
    final_r = sum(m['rating'] * m['rounds'] for m in final_matches)
    final_w = sum(m['rounds'] for m in final_matches)
    final_avg = round(final_r / final_w, 2) if final_w > 0 else "-"

    # Debug logs for summary
    print(f"  Summary Metrics: WinAvg={win_rating_avg}, Combined={weighted_combined}")

    # 4. Map History Traversal (Reversed for Chronological Order)
    map_history = []
    print(f"  Target Player: {target_name} (ID: {target_player_id})")
    for m in reversed(all_matches):
        print(f"  Fetching rank for map vs {m['opp']}...")
        rank = get_map_rank(driver, m['url'], target_player_id)
        # Attempt to capture the specific rating from traversal as verification
        map_history.append({
            'opp': m['opp'],
            'res': "Win" if m['is_win'] else "Loss",
            'rating': m['rating'],
            'rank': rank
        })

    return win_rating_avg, final_avg, weighted_combined, map_history

def main():
    if not os.path.exists(os.path.dirname(OUTPUT_PATH)):
        os.makedirs(os.path.dirname(OUTPUT_PATH))
    
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    driver = uc.Chrome(version_main=146)
    
    try:
        # Initial visit to clear bot checks/cookies

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        for url in PLAYER_URL_LIST:
            match = re.search(r'/players/(\d+)/([^?]+)', url)
            p_name = match.group(2) if match else "unknown"
            event_match = re.search(r'event=(\d+)', url)
            e_id = event_match.group(1) if event_match else "0"
            
            sheet_title = f"{p_name}_{e_id}"[:30]
            ws = wb.create_sheet(title=sheet_title)
            print(f"Scraping: {p_name} (Event {e_id})...")
            
            

            # Headers
            ws.merge_cells('A1:A2')
            ws['A1'] = "Player"
            ws.merge_cells('B1:E1')
            ws['B1'] = "Overall"
            ws.merge_cells('F1:I1')
            ws['F1'] = "Playoffs"
            ws.merge_cells('J1:M1')
            ws['J1'] = "Grand Final"
            ws.merge_cells('N1:O1')
            ws['N1'] = "Calculated Metrics"

            
            sub_headers = ["Maps", "Rating 2.0", "KPRW", "ADRW"]
            for i, sh in enumerate(sub_headers):
                for start_col in [2, 6, 10]:
                    ws.cell(row=2, column=start_col+i, value=sh)
            
            ws.cell(row=2, column=14, value="Win Map Rating")
            ws.cell(row=2, column=15, value="Combined Weighted")
            
            for row_idx in [1, 2]:
                for cell in ws[row_idx]:
                    cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center", vertical="center")
                    cell.border = Border(top=Side(style="thin", color="FFFFFF"), left=Side(style="thin", color="FFFFFF"), 
                                       right=Side(style="thin", color="FFFFFF"), bottom=Side(style="thin", color="FFFFFF"))

            # Scrape Data - Extract player name from first call
            overall_result = get_stats_from_summary(driver, url)
            scraped_name = overall_result[0]  # Player name from HLTV (preserves case)
            overall = overall_result[1:]  # (maps, rating, kprw, adrw)
            
            playoffs_result = get_stats_from_summary(driver, url, "PLAYOFFS")
            playoffs = playoffs_result[1:]
            
            final_result = get_stats_from_summary(driver, url, "GRAND_FINAL")
            final_stats = final_result[1:]
            
            win_r, final_r, combined_w, map_history = get_detailed_map_stats(driver, url)
            
            # Use scraped name if available, else fallback to URL name
            display_name = scraped_name if scraped_name else p_name
            
            # Rename sheet to use proper name
            ws.title = f"{display_name}_{e_id}"[:30]
            
            # Row 3 Summary - Use display_name (preserving case from HLTV)
            ws.append([display_name, *overall, *playoffs, *final_stats, win_r, combined_w])
            for cell in ws[3]:
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))

            # Row 5+ Map Details
            ws.append([]) # Gap
            detail_headers = ["Opponent", "Result", "Player Rating", "Rank (1-10)"]
            ws.append(detail_headers)
            header_row_idx = 5
            for i, h in enumerate(detail_headers):
                cell = ws.cell(row=header_row_idx, column=i+1)
                cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")
                cell.border = Border(top=Side(style="thin", color="FFFFFF"), left=Side(style="thin", color="FFFFFF"), 
                                   right=Side(style="thin", color="FFFFFF"), bottom=Side(style="thin", color="FFFFFF"))

            for m in map_history:
                ws.append([m['opp'], m['res'], m['rating'], m['rank']])
                for cell in ws[ws.max_row]:
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))

            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            wb.save(OUTPUT_PATH)
            print(f"Saved {p_name} progress.")

        print(f"\nCompleted! Final file: {OUTPUT_PATH}")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
