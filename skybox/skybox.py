import re 
from PIL import Image
import pytesseract
import os
from openpyxl import Workbook 
import undetected_chromedriver as uc
import time
from openpyxl import Workbook
from openpyxl import load_workbook
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
playerwebfront = "https://edge.skybox.gg/u/player-stats/"
totalwebfront = "https://edge.skybox.gg/u/leaderboards?before=01-09-2025&min_rank=&min_rounds_played=1000&order_by=HLTV_RATING2&order_by_direction=desc&since=01-01-2025" 
events_ID=[
"&events=BLAST+Bounty+2025+Season+1",
"&events=BLAST+Bounty+2025+Season+1+Finals",
"&events=IEM+Katowice+2025+Play-in",
"&events=IEM+Katowice+2025",
"&events=PGL+Cluj-Napoca+2025",
"&events=ESL+Pro+League+Season+21+Stage+1",
"&events=ESL+Pro+League+Season+21",
"&events=BLAST+Open+Lisbon+2025",
"&events=PGL+Bucharest+2025",
"&events=IEM+Melbourne+2025",
"&events=BLAST+Rivals+2025+Season+1",
"&events=PGL+Astana+2025",
"&events=IEM+Dallas+2025",
"&events=BLAST.tv+Austin+Major+2025+Stage+1",
"&events=BLAST.tv+Austin+Major+2025+Stage+2",
"&events=BLAST.tv+Austin+Major+2025",
"&events=FISSURE+Playground+1",
"&events=IEM+Cologne+2025",
"&events=IEM+Cologne+2025+Stage+1",
]
selected=["ZywOo","donk","m0NESY","ropz","sh1ro","NiKo","flameZ","xertioN","Spinx","torzsi","XANTARES","Senzu","910","mezii","zont1x","frozen","Wicadia","TeSeS","woxic","b1t","apEX","Jimpphat","EliGE","HeavyGod"]
wb = Workbook()
ws = wb.active
driver = uc.Chrome() 
names = []
player_IDs = []
file_path = os.path.join("C:\\Users\\10725\\Desktop\\hltv\\skybox", "skybox.xlsx")
events_total = ""
for i in events_ID:
    events_total += i

player_data = dict()

if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始
        name = row[0]
        mapcount = row[1]
        player_data[name] = list(row)  # 将整行数据存储为列表
else:
    wb = Workbook()
    ws = wb.active

# 提取指定区域的数字（如 ADR 值）
def extract_adr_from_region(image, box):
    region = image.crop(box)
    text = pytesseract.image_to_string(region, config='--psm 11 --oem 3 tessedit_char_whitelist=0123456789')
    match = re.search(r'(\d+)', text)
    return int(match.group(1)) if match else None



driver.get(totalwebfront+events_total+"&page="+"1")
time.sleep(30) 
content = driver.page_source 
stats_temp = re.findall("<img src=\"https://assets.skybox.gg/player-images/(.*?)\" class=",content)
for i in stats_temp:
    player_IDs.append(i.split(".", 1)[0].strip())
    names.append((i.split("\"", 1)[-1].strip()).split("\"", 1)[-1].strip())

driver.get(totalwebfront+events_total+"&page="+"2")
time.sleep(3)
content = driver.page_source 
stats_temp = re.findall("<img src=\"https://assets.skybox.gg/player-images/(.*?)\" class=",content)
for i in stats_temp:
    player_IDs.append(i.split(".", 1)[0].strip())
    names.append((i.split("\"", 1)[-1].strip()).split("\"", 1)[-1].strip())

driver.get(totalwebfront+events_total+"&page="+"3")
time.sleep(5)
content = driver.page_source 
stats_temp = re.findall("<img src=\"https://assets.skybox.gg/player-images/(.*?)\" class=",content)
for i in stats_temp:
    player_IDs.append(i.split(".", 1)[0].strip())
    names.append((i.split("\"", 1)[-1].strip()).split("\"", 1)[-1].strip())

driver.get(totalwebfront+events_total+"&page="+"4")
time.sleep(5)
content = driver.page_source 
stats_temp = re.findall("<img src=\"https://assets.skybox.gg/player-images/(.*?)\" class=",content)
for i in stats_temp:
    player_IDs.append(i.split(".", 1)[0].strip())
    names.append((i.split("\"", 1)[-1].strip()).split("\"", 1)[-1].strip())

for name,player_ID in zip(names,player_IDs):
    if name in player_data:
        continue
    if name not in selected:
        continue
    row=[name,player_ID]
    rounds = 0.0
    Full_Buy_vs_Full_Buy_Rating = 0.0
    Full_Buy_vs_Half_Eco_Rating = 0.0
    Pistol_rounds_Rating = 0.0
    FBFB_Adr = 0.0
    PTPT_Adr = 0.0
    FBHB_Adr = 0.0
    FBEco_Adr = 0.0
    HBFB_Adr = 0.0
    EcoFB_Adr = 0.0
    for event_ID in events_ID:
        driver.get(playerwebfront+player_ID+"/intro?min_rank=&since=01-01-2025&map=all&side=both"+event_ID)
        time.sleep(5)
        content = driver.page_source 
        rounds_stats = re.findall("<span class=\"text-base\">(.*?)</span>",content)
        if rounds_stats[1] == "0":
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            row.append(0.0)
            continue
        rounds_event = float(rounds_stats[1])
        rounds += rounds_event
        row.append(rounds_event)
        stats_part1=re.findall("<span class=\"text-2xl\">(.*?)</span>",content)
        Full_Buy_vs_Full_Buy_Rating += rounds_event*float(stats_part1[2])
        Full_Buy_vs_Half_Eco_Rating += rounds_event*float(stats_part1[3])
        Pistol_rounds_Rating += rounds_event*float(stats_part1[4])
        row.append(float(stats_part1[2]))
        row.append(float(stats_part1[3]))
        row.append(float(stats_part1[4]))
        driver.get(playerwebfront+player_ID+"/buytypes?min_rank=&since=01-01-2025&map=all&side=both"+event_ID)
        content = driver.page_source 
        time.sleep(5)
        canvas = driver.find_element("tag name", "canvas")
        driver.execute_script("arguments[0].scrollIntoView(true);", canvas)
        time.sleep(0.5)
        canvas.screenshot("buytypes_canvas.png")
        image_path = "buytypes_canvas.png"
        img = Image.open(image_path)
        # 自定义每个柱状图的数字区域（左，上，右，下）→ 根据图像大小可调整
        regions = {
            "FB V FB": (50, 0, 180, 616),
            "Pistol V Pistol": (200, 0, 300, 616),
            "FB V HB": (320, 0, 460, 616),
            "FB V Eco": (480, 0, 600, 616),
            "HB V FB": (600, 0, 750, 616),
            "Eco V FB": (750, 0, 900, 616),
        }
        results = []
        for label, box in regions.items():
            adr = extract_adr_from_region(img, box)
            if adr == None:
                results.append("0")
                continue
            results.append(adr)
        FBFB_Adr += rounds_event*float(results[0])
        PTPT_Adr += rounds_event*float(results[1])
        FBHB_Adr += rounds_event*float(results[2])
        FBEco_Adr += rounds_event*float(results[3])
        HBFB_Adr += rounds_event*float(results[4])  
        EcoFB_Adr += rounds_event*float(results[5])
        row.append(float(results[0]))
        row.append(float(results[1]))
        row.append(float(results[2]))
        row.append(float(results[3]))
        row.append(float(results[4]))
        row.append(float(results[5]))
    row.append(Full_Buy_vs_Full_Buy_Rating / rounds)
    row.append(Full_Buy_vs_Half_Eco_Rating/ rounds)
    row.append(Pistol_rounds_Rating / rounds)
    row.append(FBFB_Adr / rounds)
    row.append(PTPT_Adr / rounds)
    row.append(FBHB_Adr / rounds)
    row.append(FBEco_Adr / rounds)
    row.append(HBFB_Adr / rounds)
    row.append(EcoFB_Adr / rounds) 
    row.append("")
    print(row)
    ws.append(row)
    wb.save(file_path)

""" 
rui013@e.ntu.edu.sg
Shaol1n@kungfu
"""
