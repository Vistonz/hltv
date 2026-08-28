import re
import os
import time
from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc

# ================= 配置区域 =================
# 1. 赛事过滤器配置
# eventfilter: HLTV 统计页面的事件 ID 过滤参数
# eventnamefilter: 对应的中文赛事名称，将作为 Excel 中的 Event_ID
eventfilter = [
    "&event=8246",  "&event=8240", "&event=8047",
    "&event=8413", "&event=8248","&event=8048","&event=8242","&event=8250","&event=8049","&event=8243","&event=8263","&event=8301"
]
eventnamefilter = [
    "BLAST赏金赛S1",  "IEM克拉科夫", "PGL克卢日纳波卡",
    "EPL S23", "BLAST鹿特丹公开赛","PGL布加勒斯特","IEM里约","BLAST对抗赛S1","PGL阿斯塔纳","IEM亚特兰大","CAC","IEM科隆Major"
]

# 2. 输出路径配置
output_dir = r"/home/hongbin/Desktop/hltv/hltv top/2026 S1"
grade_filename = "grade.xlsx"

# 3. 爬虫性能配置
sleep_first = 25.6657    # 第一次加载页面的等待时间（用于过验证）
sleep_others = 0.6657   # 后续操作的等待时间
chrome_version = 148   # Chrome 浏览器主版本号

# 4. 其他配置
keyword = ">"
# ===========================================


def scrape_season_events(
    eventfilter=eventfilter,
    eventnamefilter=eventnamefilter,
    output_dir=output_dir,
    grade_filename=grade_filename,
    sleep_first=sleep_first,
    sleep_others=sleep_others,
    chrome_version=chrome_version,
    keyword=keyword,
):
    """HLTV 赛季选手数据爬虫主入口: 按赛事列表抓选手/队伍/排名, 汇总写 grade.xlsx.

    参数均可传入覆盖 (默认值 = 脚本内硬编码配置, 行为不变):
      eventfilter / eventnamefilter  赛事 event id 与中文名列表 (一一对应)
      output_dir / grade_filename    输出目录与文件名
      sleep_first / sleep_others     首次/后续页面等待秒数
      chrome_version                 Chrome 主版本号 (原值 148; 本机 151 需传 151)
      keyword                        清洗 name/rating 用分隔符
    """
    file_path = os.path.join(output_dir, grade_filename)

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # === 初始化浏览器 ===
    driver = uc.Chrome(version_main=chrome_version)

    existing_events = set()

    # === 读取现有文件或创建新文件 ===
    if os.path.exists(file_path):
        print(f"发现现有文件：{file_path}，正在读取已存在的赛事...")
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1]:
                    existing_events.add(row[1])
            print(f"已跳过的赛事列表: {list(existing_events)}")
        except Exception as e:
            print(f"读取文件出错，将重新创建: {e}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Player Stats"
            ws.append([
                "Player", "Event_ID", "Country","CountryURL","Team", "Grade", "Rating", "Team_Rank",
                "Rating_Diff(%)", "KPR", "ADR", "DPR", "RS", "KAST", "eaKPR", "eaADR", "eaDPR", "eaMK", "eaKAST",
            ])
    else:
        print(f"未发现针对当前赛季的文件，在目标路径创建新文件: {file_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Player Stats"
        ws.append([
            "Player", "Event_ID", "Country","CountryURL","Team", "Grade", "Rating", "Team_Rank",
            "Rating_Diff(%)", "KPR", "ADR", "DPR", "RS", "KAST", "eaKPR", "eaADR", "eaDPR", "eaMK", "eaKAST",
        ])

    # === 开始抓取数据 ===
    webfront = "https://www.hltv.org/stats/players?startDate=all&minMapCount=0"
    teamfront = "https://www.hltv.org/stats/teams?startDate=all&minMapCount=0"

    player_stats = {}
    zz = 0
    j = 0

    for i in eventfilter:
        current_event_name = eventnamefilter[j]

        # === 检查跳过逻辑 ===
        if current_event_name in existing_events:
            print(f"赛事 '{current_event_name}' 已存在于表格中，跳过...")
            j += 1
            continue

        print(f"开始抓取赛事: {current_event_name}...")
        event_id = current_event_name
        j += 1

        # 1. 抓取 Player Stats
        driver.get(webfront + i)
        time.sleep(sleep_first if zz == 0 else sleep_others)
        zz = 1
        content = driver.page_source

        Name1 = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>', content)
        Mapcount1 = re.findall('<td class="statsDetail">(.*?)</td>', content)
        Team = re.findall('class="teamCol" data-sort="(.*?)"><', content)
        Rating = re.findall('class="ratingCol(.*?)</td>', content)
        Id = re.findall('<a href="/stats/players(.*?)" data-tooltip-id="uniqueTooltipId', content)
        Rounds = re.findall('<td class="statsDetail gtSmartphone-only">(.*?)</td>', content)
        Flag = re.findall('class="flag" title="(.*?)">', content)
        FlagURL = re.findall('" src="(.*?)" class="flag"',content)

        # 2. 抓取 Team Rating
        team_rating_per_event = {}
        driver.get(teamfront + i)
        time.sleep(sleep_others)
        content = driver.page_source
        Team1 = re.findall('data-tooltip-id="uniqueTooltipId-(.*?)</a></td>', content)
        Rating1 = re.findall('class="ratingCol(.*?)</td>', content)
        team_rating = {}
        for idx in range(len(Team1)):
            team_name = Team1[idx].split(keyword, 1)[-1].strip()
            team_rating_value = Rating1[idx].split(keyword, 1)[-1].strip()
            try:
                team_rating[team_name] = float(team_rating_value)
            except:
                team_rating[team_name] = None
        team_rating_per_event[event_id] = team_rating

        # 清洗玩家数据
        Name = []
        Mapcount = []
        tempzz = 0
        while tempzz * 2 < len(Mapcount1):
            Mapcount.append(Mapcount1[tempzz * 2])
            if len(Mapcount1) == len(Name1):
                Name.append(Name1[tempzz * 2].split(keyword, 1)[-1].strip())
            else:
                Name.append(Name1[tempzz].split(keyword, 1)[-1].strip())
            Team[tempzz] = Team[tempzz]
            Rating[tempzz] = Rating[tempzz].split(keyword, 1)[-1].strip()
            tempzz += 1

        # 3. 抓取队伍排名 Grade
        team_grade_dict = {}
        extracted_event_ids = re.findall(r'event=(\d+)', i)
        print(f"正在抓取排名，检测到子赛事ID: {extracted_event_ids}")

        for eid in extracted_event_ids:
            rank_url = f"https://www.hltv.org/events/{eid}/1"
            driver.get(rank_url)
            time.sleep(sleep_others)
            content = driver.page_source
            Team2 = re.findall('"><a href="/team/(.*?)</a></div>', content)
            Grade = re.findall('</a></div>\n                      <div>(.*?)</div>', content)
            for idx in range(len(Grade)):
                if idx < len(Team2):
                    team_name = Team2[idx].split(keyword, 1)[-1].strip()
                    grade_value = Grade[idx].split(keyword, 1)[-1].strip()
                    team_grade_dict[team_name] = grade_value

        # 4. 构建选手数据
        team_to_players = {}
        for idx, name in enumerate(Name):
            player_id = Id[idx].split("/")[1] if idx < len(Id) else "unknown"
            team = Team[idx] if idx < len(Team) else "unknown"
            rating = Rating[idx] if idx < len(Rating) else "0"
            maps = Mapcount[idx] if idx < len(Mapcount) else "unknown"
            rounds = Rounds[idx] if idx < len(Rounds) else "unknown"
            flag = Flag[idx] if idx < len(Flag) else "unknown"
            flagURL = FlagURL[idx] if idx < len(FlagURL) else "unknown"

            try:
                rating_float = float(rating)
            except:
                rating_float = 0.0

            # 抓取选手详细页面数据
            statsuffix = Id[idx].replace('amp;', '')
            driver.get("https://www.hltv.org/stats/players" + statsuffix)
            time.sleep(sleep_others)
            content = driver.page_source

            Playerstatform=re.findall('<div class="player-summary-stat-box-data traditionalData">(.*?)</div>',content)
            try:
                dpr = Playerstatform[0]
                kast1=re.findall('<div class="player-summary-stat-box-data traditionalData">(.*?)<span',content)
                kast = kast1[0]
                rs1 = re.findall('<div class="player-summary-stat-box-data">(.*?)<span class="',content)
                rs = rs1[0]
                adr = Playerstatform[3]
                kpr = Playerstatform[4]
                Playerstatform_eco=re.findall('<div class="player-summary-stat-box-data ecoAdjustedData hidden">(.*?)</div>',content)
                eadpr = Playerstatform_eco[0]
                eakast1=re.findall('<div class="player-summary-stat-box-data ecoAdjustedData hidden">(.*?)<span',content)
                eakast = eakast1[0]
                eamk = Playerstatform_eco[2]
                eaadr = Playerstatform_eco[3]
                eakpr = Playerstatform_eco[4]
            except IndexError:
                dpr = kast = rs = adr = kpr = eadpr = eakast = eamk = eaadr = eakpr = "0"

            if team not in team_to_players:
                team_to_players[team] = []
            team_to_players[team].append((name, rating_float))

            if name not in player_stats:
                player_stats[name] = {}

            player_stats[name][event_id] = {
                "player_id": player_id,
                "flag": flag,
                "flagURL": flagURL,
                "team": team,
                "rating": rating,
                "rating_float": rating_float,
                "maps": maps,
                "rounds": rounds,
                "team_rating": team_rating_per_event[event_id].get(team, None),
                "rank_in_team": None,
                "rating_diff_pct": None,
                "grade": team_grade_dict.get(team, "N/A"),
                "dpr": dpr,
                "kast": kast,
                "rs": rs,
                "adr": adr,
                "kpr": kpr,
                "eadpr":eadpr,
                "eakast":eakast,
                "eamk":eamk,
                "eaadr":eaadr,
                "eakpr":eakpr,
            }

        # 计算队内排名和rating差
        for team_name, players in team_to_players.items():
            sorted_players = sorted(players, key=lambda x: x[1], reverse=True)
            for rank, (name, _) in enumerate(sorted_players, 1):
                if event_id in player_stats[name]:
                    player_stats[name][event_id]["rank_in_team"] = rank
                    team_rating_val = player_stats[name][event_id]["team_rating"]
                    if team_rating_val:
                        diff_pct = (player_stats[name][event_id]["rating_float"] - team_rating_val) / team_rating_val
                        player_stats[name][event_id]["rating_diff_pct"] = round(diff_pct, 2)

        # === 写入当前赛事的数据到Excel ===
        for player_name in Name:
            if player_name in player_stats and event_id in player_stats[player_name]:
                stats = player_stats[player_name][event_id]
                ws.append([
                    player_name,
                    event_id,
                    stats.get("flag", ""),
                    stats.get("flagURL", ""),
                    stats.get("team", ""),
                    stats.get("grade", ""),
                    float(stats.get("rating", 0)),
                    stats.get("rank_in_team", ""),
                    stats.get("rating_diff_pct", ""),
                    float(stats.get("kpr", 0)),
                    float(stats.get("adr", 0)),
                    float(stats.get("dpr", 0)),
                    stats.get("rs", ""),
                    stats.get("kast", ""),
                    float(stats.get("eakpr", 0)),
                    float(stats.get("eaadr", 0)),
                    float(stats.get("eadpr", 0)),
                    float(stats.get("eamk", 0)),
                    stats.get("eakast", ""),
                ])

        # === 保存文件 ===
        wb.save(file_path)
        print(f"赛事 {event_id} 数据抓取并保存完成。")

    driver.quit()
    print(f"所有赛事数据抓取完成，Excel 文件已保存至：{file_path}")


if __name__ == "__main__":
    scrape_season_events()
